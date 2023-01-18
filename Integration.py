import os
import sys
import numpy as np
import datetime as dt
import pandas as pd
import pandas_datareader.data as pdd
import sqlalchemy as sa
import oracledb as odb
import matplotlib.pyplot as plt
from matplotlib import style
from sklearn import svm
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor,AdaBoostRegressor
from sklearn import linear_model,tree
from sklearn.metrics import classification_report
from openpyxl import load_workbook

un="HENRY_ZHANG"
pw="Q1w2e3r4"
cs="PROD"

ms=r'C:\Users\henry.zhang\Desktop\Projects\Intergration\Mistersheep.xlsm'

def Run():
    func={'Color':color,'Data':data,'DM':dm,'Script':script,'Suggest':machinelearning}

    wb = xw.Book.caller()
    ws=wb.sheets['Setup']

    func[ws['I6'].value]()


def color():
    
    #Call book
    wb = xw.Book.caller()
    ws=wb.sheets['Setup']
    
    #Credentials
    un=ws['J1'].value
    pw=ws['J2'].value
    cd=ws['N1'].value

    #Locate variables in 'Setup' sheet
    cs="PROD"
    keep=ws['I10'].value
    lag=ws['I9'].value

    #Confirm Date format
    if ws['I11'].value=='US':
        formate='%m/%d/%Y'
    elif ws['I11'].value=='EU':
        formate='%d/%m/%Y'
    
    #Establishing connection to Database
    con=odb.connect(user=un,password=pw,dsn=cs)

    colorqry="""
    WITH SEC_MASTER AS (
        SELECT MASTER.SEC_ID, MASTER.SEC_ID_TYPE , FIELD_NAME,
            CASE
                WHEN FIELD_TYPE = 'String' THEN FIELD_STRING_VALUE
                WHEN FIELD_TYPE = 'Integer' THEN TO_CHAR(FIELD_INTEGER_VALUE)
            END AS FIELD_VALUE
        FROM SECURITY_MASTER MASTER
        LEFT JOIN SEC_INDICATIVE_INFO DEAL
            ON DEAL.SEC_ID = MASTER.SEC_ID
            AND DEAL.SEC_ID_TYPE = MASTER.SEC_ID_TYPE
            AND DEAL.FIELD_NAME in( 'DealType','Currency','OriginalBestRatingScore','MkTicker')

        WHERE prc_config_sec_type = 'CLO'
            AND prc_CONFIG_KEY_NAME NOT IN ('All','CDO','CDO_Flat', 'TruPS','CRT')
            and sec_status='Approved'
    ), CLO_LIST AS (
        SELECT DISTINCT ROOTID,
            NVL(LIST.CUSIP,  INTEX.CUSIP) CUSIP,
            NVL(LIST.ISIN,  INTEX.ISIN) ISIN,
            REGEXP_SUBSTR(BBG, '^([A-Z0-9]+) ([A-Z0-9-]+) ([A-Z0-9-]+)$' ,1,1,'i',1) SHELF,
            REGEXP_SUBSTR(BBG, '^([A-Z0-9]+) ([A-Z0-9-]+) ([A-Z0-9-]+)$' ,1,1,'i',2) SERIES,
            REGEXP_SUBSTR(BBG, '^([A-Z0-9]+) ([A-Z0-9-]+) ([A-Z0-9-]+)$' ,1,1,'i',3) CLASS
        FROM (
            SELECT
                CONNECT_BY_ROOT ORIGID ROOTID,
                CASE DUPLICATETYPE WHEN 'CUSIP' THEN DUPLICATEDID ELSE '' END CUSIP,
                CASE DUPLICATETYPE WHEN 'ISIN'  THEN DUPLICATEDID ELSE '' END ISIN
            FROM IDENTIFIERMAPPINGS
            START WITH ORIGID IN (SELECT SEC_ID FROM SEC_MASTER )
            CONNECT BY NOCYCLE ORIGID = PRIOR DUPLICATEDID

            UNION

            SELECT SEC_ID ROOTID,
                CASE SEC_ID_TYPE WHEN 'CUSIP' THEN SEC_ID ELSE '' END CUSIP,
                CASE SEC_ID_TYPE WHEN 'ISIN'  THEN SEC_ID ELSE '' END ISIN
            FROM SEC_MASTER
            ) LIST

        LEFT JOIN INTEXBONDINFO INTEX
        ON INTEX.CUSIP = ROOTID

        LEFT JOIN (
            SELECT SEC_ID, FIELD_VALUE BBG
            FROM SEC_MASTER
            WHERE FIELD_NAME = 'MkTicker'
        )INFO
        ON INFO.SEC_ID = ROOTID
    ),
        TAINTED_COLOR as (
        SELECT
            QUOTES.ROWID QUOTEID, MESSAGE_ID,
            CONTRIBUTOR, OWNER,
            PRICE_BID, PRICE_ASK,
            SPREAD_BID, SPREAD_ASK,
            BENCHMARK, WAL, PRICELEVEL,
            CONFIDENCE,
            CLO_LIST.ROOTID ROOT_CUSIP, QUOTES.CUSIP, QUOTES.ISIN,
            QUOTES.SHELF, QUOTES.SERIES, QUOTES.CLASS,
            NVL(BIAS, NVL2(PRICE_BID,'BID','OFFER')) BIAS,
            CASE WHEN CONTRIBUTOR  in ('ED&'||'F', 'Aurul')THEN 1 ELSE 0 END TAINTED,
            TIME,TRUNC(TIME) MKT_DATE,
            NVL(PRICE_BID,PRICE_ASK) PRICE
        FROM SF_QUOTES_ABS QUOTES

        RIGHT JOIN CLO_LIST
            ON CLO_LIST.CUSIP = QUOTES.CUSIP
            OR CLO_LIST.ISIN = QUOTES.ISIN
            OR (CLO_LIST.SHELF = QUOTES.SHELF
            AND CLO_LIST.SERIES = QUOTES.SERIES
            AND CLO_LIST.CLASS = QUOTES.CLASS)

        WHERE
            OWNER <> 'DEMO'
            AND QUOTES.TIME <=  trunc(sysdate)+1--#DATE  +1
            AND QUOTES.TIME >=  trunc(sysdate)-"""+str(lag)+"""--#DATE
            AND ( PRICE_BID IS NOT NULL OR PRICE_ASK IS NOT NULL)
            and contributor not in ('noreply@solveadvisors.com','State Street ETF Group')
    ),
    FILTERED_COLOR AS (
        SELECT *
        FROM (
            SELECT COLOR.*,
            ROW_NUMBER() OVER ( PARTITION BY ROOT_CUSIP, MKT_DATE ORDER BY TIME ASC ) RECENT
            FROM (
                SELECT COLOR.*,
                MEDIAN(PRICE) OVER ( PARTITION BY ROOT_CUSIP, MKT_DATE ) MEDIAN_PRICE
                FROM (
                    SELECT COLOR.*,
                    RANK() OVER (PARTITION BY ROOT_CUSIP, MKT_DATE ORDER BY DECODE(
                    BIAS,'TRADE',1,'TRADE CONFIRM',1,'BWIC COVER',2,'BID',3,'MARKET',3,'BUYER',3,
                    'BWIC REO',5,'BWIC TALK',5,'BWIC',5,'OFFER',6,'VALUATION',7,10
                    ) ASC
                    ) BIAS_RANK
                    FROM (
                        SELECT TAINTED_COLOR.*,
                        RANK() OVER (PARTITION BY ROOT_CUSIP, TRUNC(TIME) ORDER BY CONFIDENCE DESC)
                        + CASE WHEN TAINTED = 1 THEN 1 ELSE 0 END CONFIDENCE_RANK
                        FROM TAINTED_COLOR
                    )COLOR
                    WHERE CONFIDENCE_RANK = 1 OR CONFIDENCE >= 8
                ) COLOR
                WHERE BIAS_RANK = 1
            ) COLOR
            WHERE MEDIAN_PRICE = PRICE
        ) COLOR
        WHERE RECENT = 1
    ),
    COLOUR as (
        SELECT
            SYSDATE TIME_QUERIED,
            TIME TIME_RECEIVED,
            TO_CHAR(MESSAGE_ID) MESSAGE,
            ROOT_CUSIP, CUSIP CUSIP_COLOR,
            ISIN ISIN_COLOR,SHELF || ' ' || SERIES || ' ' || CLASS TICKER,
            NVL(RTG.FIELD_VALUE, '100') ORIG_BEST_RATING,
            NVL(CURR.FIELD_VALUE, '') CURRENCY,
            PRICE_BID, PRICE_ASK, SPREAD_BID, SPREAD_ASK,
            BENCHMARK, WAL, PRICELEVEL,DECODE(BIAS, 'MARKET','BID','BUYER','BID',NULL,NVL2(PRICE_BID,'BID','OFFER'),BIAS) BIAS,
            CONTRIBUTOR, OWNER,
            NVL(PRICE_BID, PRICE_ASK) PRICE, CONFIDENCE, QUOTEID

        FROM FILTERED_COLOR

        LEFT JOIN SEC_MASTER RTG
            ON RTG.SEC_ID = ROOT_CUSIP
            AND RTG.FIELD_NAME = 'OriginalBestRatingScore'

        LEFT JOIN SEC_MASTER CURR
            ON CURR.SEC_ID = ROOT_CUSIP
            AND CURR.FIELD_NAME = 'Currency'
    ),
    ACCEPTED as (
        select root_cusip,
            sm.prc_config_key_name,
            CASE
                WHEN PRICE > 105 AND sm.prc_config_key_name in ('1.0_Mezz', '2.0_Mezz', '2.0_Senior') THEN 1
                WHEN PRICE < 10 AND sm.prc_config_key_name in ('1.0_Mezz', '2.0_Mezz', '2.0_Senior') THEN 1
                ELSE 0
            END as REJECTED


        from COLOUR col,security_master sm

        where col.root_cusip=sm.sec_id
            and sm.sec_status='Approved'
            and sm.prc_config_sec_type = 'CLO'
    )




    select

        to_char(col.message),
        trunc(col.TIME_RECEIVED) date1,
        null ,
        null ,
        null ,
        col.ticker,
        col.root_cusip,
        trunc(sysdate),
        nvl(col.price,0) PRICE_LEVEL,
        nvl(col.price_bid,0) Bid,
        nvl(col.price_ask,0) Ask,
        nvl(col.price,0) PRICE_LEVEL,
        nvl(col.price,0) Px,
        --case when (nvl(col.price_ask,0)!=0 and col.bias='BID') then (nvl(col.price_bid,0)+nvl(col.price_ask,0))/2 else nvl(col.price,0) end as Px,
        case when (nvl(col.price_ask,0)!=0 and col.bias='BID') then 'MARKET' else col.bias end as BIAS,
        col.contributor,
        DECODE(BIAS, 'TRADE CONFIRM', 1, 'BWIC COVER',2, 'BWIC REO', 3, 'MARKET', 3, 'BUYER', 3,'BID',3, 'COLOR', 3, 'OFFER', 4, 'BWIC TALK', 5, 'VALUATION', 6) RANK

    from COLOUR col, ACCEPTED a
    where col.root_cusip=a.root_cusip
        and a.rejected=0
        and currency='USD'

    order by RANK,date1 desc,Px desc
    """
    #Read old colors
    df_oc=pd.read_excel(cd,sheet_name='Colours')

    #Execute color query and save in dataframe
    df_nc=pd.read_sql(colorqry,con)

    #Align column names for merging, adding str and \t to avoid scientific notation and thus loss of precision
    df_nc.columns = df_oc.columns

    #Clean new colors
    df_nc.sort_values(by=['Rank','DATE'],ascending=[True,False],inplace=True)
    df_nc.drop_duplicates(subset='Cusip',keep='first',inplace=True)
    df_nc.reset_index(drop=True,inplace=True)

    #Change dtpye to date for later addition to check if dates are neighbores
    df_nc["Date2"] =  pd.to_datetime(df_nc["Date2"], format=formate)
    df_oc["Date2"] =  pd.to_datetime(df_oc["Date2"], format=formate)

    #Prepare to check if New color has past record in Comments
    df=df_oc['Cusip'].value_counts()

    #get column names
    prs="Px"
    dat="Date2"
    rak="Rank"
    csp="Cusip"

    #Check for stale in the new colors
    for i in range(0,df_nc.shape[0]):
        cusip=df_nc.loc[i,csp]

        if cusip in df and df[cusip]>=keep:
            idx=df_oc.index[df_oc.Cusip==cusip].tolist()[0]

            if df_nc.loc[i,prs]==df_oc.loc[idx,prs] and df_nc.loc[i,dat] in {df_oc.loc[idx,dat], df_oc.loc[idx,dat]+timedelta(lag)} and df_nc.loc[i,rak]==df_oc.loc[idx,rak]:
                df_oc.drop(index=idx,inplace=True)


    #Combine colors
    dfc=pd.concat([df_nc, df_oc], axis=0, ignore_index=True)

    #Clean
    dfc.sort_values(by=['Date2','Rank','DATE'],ascending=[False,True,False],inplace=True)
    dfct=dfc.groupby('Cusip').head(keep)

    print(dfct.head(3))

    #Setup writer for date_format
    #Writing it back in
    with pd.ExcelWriter(cd,date_format='YYY/MM/DD') as writer:
        dfct.to_excel(writer,sheet_name='Colours',index=False)

    #Close Comments

    #Closing Connection
    con.close()


###################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################
###################################################################################################################################################################

def data():

    wb = xw.Book.caller()
    ws=wb.sheets['Setup']

    #Credentials
    un=ws['J1'].value
    pw=ws['J2'].value
    cd=ws['N1'].value

    #Locate variables in 'Setup' sheet
    wsm=wb.sheets['Main']
    cs="PROD"

    sectors=ws['I13'].value
    batch=ws['I5'].value
    dat=ws['I7'].value
    cd=ws['N1'].value

    #Get fields
    #Two "down" in df1 representing Excel shortcut ctrl+shift+down twice, meaning blank translation (custom fields) must be at top or wont be selected
    df1 = ws['A1:B1'].expand('down').expand("down").options(pd.DataFrame,index=False).value
    df2 = ws['C:C'].expand('down').options(pd.DataFrame,index=False).value

    header1=df1.columns.values[0]
    header2=df2.columns.values[0]
    header3=df1.columns.values[1]

    all_column_names = df2[header2].values.tolist()
    #get message id index for future fomatting
    #x=all_column_names.index('Message ID')
    #mesid=string.ascii_lowercase[x].upper()

    #Vlookup
    df1.rename(columns={header1:header2},inplace=True)
    dff=pd.merge(df2,df1,how='left',on=header2)

    #Drop the custom fields and feed the rest into SQL
    dff.dropna(inplace=True)

    sql_column_names=dff[header2].values.tolist()

    field = dff[header3].values.tolist()
    fq= ', '.join(field) 

    #Confirm Date format
    if ws['I11'].value=='US':
        formate='%m/%d/%Y'
    elif ws['I11'].value=='EU':
        formate='%d/%m/%Y'

    #Establishing connection to Database
    con=odb.connect(user=un,password=pw,dsn=cs)

    #Master query from Chris
    masterqry="""
    Select """+fq+"""

    From Security_master sm,Pricing_Results pr
    left join (select sec_id,
    max(case when field_name = 'AccumUnrealizedWritedown' then FIELD_Float_VALUE else null end) AccumUnrealizedWritedown,
    max(case when field_name = 'AssetCount' then FIELD_Integer_VALUE else null end) AssetCount,
    max(case when field_name = 'AssetCountWithAvailablePrice' then FIELD_Integer_VALUE else null end) AssetCountWithAvailablePrice,
    max(case when field_name = 'AssetCountWithNoAvailablePrice' then FIELD_Integer_VALUE else null end) AssetCountWithNoAvailablePrice,
    max(case when field_name = 'AverageLife' then FIELD_Float_VALUE else null end) AverageLife,
    max(case when field_name = 'CallDate' then FIELD_Integer_VALUE else null end) CallDate,
    max(case when field_name = 'CapitalEquipment' then FIELD_Float_VALUE else null end) CapitalEquipment,
    max(case when field_name = 'CCCAmount' then FIELD_Float_VALUE else null end) CCCAmount,
    max(case when field_name = 'CollateralManager' then FIELD_String_VALUE else null end) CollateralManager,
    max(case when field_name = 'CollateralManagerTier' then FIELD_String_VALUE else null end) CollateralManagerTier,
    max(case when field_name = 'CollateralType' then FIELD_String_VALUE else null end) CollateralType,
    max(case when field_name = 'Coupon' then FIELD_Float_VALUE else null end) Coupon,
    max(case when field_name = 'CoverageRatio' then FIELD_Float_VALUE else null end) CoverageRatio,
    max(case when field_name = 'CovLiteBalance' then FIELD_Float_VALUE else null end) CovLiteBalance,
    max(case when field_name = 'CovLiteBalancePercent' then FIELD_Float_VALUE else null end) CovLiteBalancePercent,
    max(case when field_name = 'Currency' then FIELD_String_VALUE else null end) Currency,
    max(case when field_name = 'CurrentCollatBalance' then FIELD_Float_VALUE else null end) CurrentCollatBalance,
    max(case when field_name = 'CurrentFitchRating' then FIELD_String_VALUE else null end) CurrentFitchRating,
    max(case when field_name = 'CurrentMoodysRating' then FIELD_String_VALUE else null end) CurrentMoodysRating,
    max(case when field_name = 'CurrentSPRating' then FIELD_String_VALUE else null end) CurrentSPRating,
    max(case when field_name = 'CurrentTrancheBalance' then FIELD_Float_VALUE else null end) CurrentTrancheBalance,
    max(case when field_name = 'Cusip' then FIELD_String_VALUE else null end) Cusip,
    max(case when field_name = 'DealName' then FIELD_String_VALUE else null end) DealName,
    max(case when field_name = 'DealType' then FIELD_String_VALUE else null end) DealType,
    max(case when field_name = 'EquityBalance' then FIELD_Float_VALUE else null end) EquityBalance,
    max(case when field_name = 'EquityPrice' then FIELD_Float_VALUE else null end) EquityPrice,
    max(case when field_name = 'FinalNAV' then FIELD_Float_VALUE else null end) FinalNAV,
    max(case when field_name = 'FitchWarfCurrent' then FIELD_Float_VALUE else null end) FitchWarfCurrent,
    max(case when field_name = 'FitchWarfLimit' then FIELD_Float_VALUE else null end) FitchWarfLimit,
    max(case when field_name = 'FloaterCap' then FIELD_Float_VALUE else null end) FloaterCap,
    max(case when field_name = 'FloaterFloor' then FIELD_Float_VALUE else null end) FloaterFloor,
    max(case when field_name = 'FloaterIndex' then FIELD_String_VALUE else null end) FloaterIndex,
    max(case when field_name = 'FloaterMargin' then FIELD_Float_VALUE else null end) FloaterMargin,
    max(case when field_name = 'FloaterMultiplier' then FIELD_Float_VALUE else null end) FloaterMultiplier,
    max(case when field_name = 'ForestProducts' then FIELD_Float_VALUE else null end) ForestProducts,
    max(case when field_name = 'IncentiveMgmtFeeCurrPaid' then FIELD_Float_VALUE else null end) IncentiveMgmtFeeCurrPaid,
    max(case when field_name = 'Insurer' then FIELD_String_VALUE else null end) Insurer,
    max(case when field_name = 'InterestCollectionAccount' then FIELD_Float_VALUE else null end) InterestCollectionAccount,
    max(case when field_name = 'InterestType' then FIELD_String_VALUE else null end) InterestType,
    max(case when field_name = 'IO_EquityBalance' then FIELD_Float_VALUE else null end) IO_EquityBalance,
    max(case when field_name = 'IsCallImpacted' then FIELD_String_VALUE else null end) IsCallImpacted,
    max(case when field_name = 'Issuer' then FIELD_String_VALUE else null end) Issuer,
    max(case when field_name = 'IssuerDate' then FIELD_String_VALUE else null end) IssuerDate,
    max(case when field_name = 'LastPaymentDate' then FIELD_Integer_VALUE else null end) LastPaymentDate,
    max(case when field_name = 'LeveragedRatio' then FIELD_Float_VALUE else null end) LeveragedRatio,
    max(case when field_name = 'LiquidityScore' then FIELD_Integer_VALUE else null end) LiquidityScore,
    max(case when field_name = 'LoanCoverageCount' then FIELD_Integer_VALUE else null end) LoanCoverageCount,
    max(case when field_name = 'LoansSubTypeBalancePercent' then FIELD_Float_VALUE else null end) LoansSubTypeBalancePercent,
    max(case when field_name = 'MarketValue' then FIELD_Float_VALUE else null end) MarketValue,
    max(case when field_name = 'MarketValueOC' then FIELD_Float_VALUE else null end) MarketValueOC,
    max(case when field_name = 'MkTicker' then FIELD_String_VALUE else null end) MkTicker,
    max(case when field_name = 'MLPEvaluatedCoveragePercent' then FIELD_Float_VALUE else null end) MLPEvaluatedCoveragePercent,
    max(case when field_name = 'MonthTillNextPayment' then FIELD_Integer_VALUE else null end) MonthTillNextPayment,
    max(case when field_name = 'MoodysWarfCurrent' then FIELD_Float_VALUE else null end) MoodysWarfCurrent,
    max(case when field_name = 'MoodysWarfLimit' then FIELD_Float_VALUE else null end) MoodysWarfLimit,
    max(case when field_name = 'MVCollateral' then FIELD_Float_VALUE else null end) MVCollateral,
    max(case when field_name = 'NAVRatio' then FIELD_Float_VALUE else null end) NAVRatio,
    max(case when field_name = 'NoCountryListed' then FIELD_Float_VALUE else null end) NoCountryListed,
    max(case when field_name = 'NotReportedIndustry' then FIELD_Float_VALUE else null end) NotReportedIndustry,
    max(case when field_name = 'OptionRedemption' then FIELD_String_VALUE else null end) OptionRedemption,
    max(case when field_name = 'OriginalBestRatingScore' then FIELD_Integer_VALUE else null end) OriginalBestRatingScore,
    max(case when field_name = 'OriginalFitchRating' then FIELD_String_VALUE else null end) OriginalFitchRating,
    max(case when field_name = 'OriginalFitchScore' then FIELD_Integer_VALUE else null end) OriginalFitchScore,
    max(case when field_name = 'OriginalMoodysRating' then FIELD_String_VALUE else null end) OriginalMoodysRating,
    max(case when field_name = 'OriginalMoodysScore' then FIELD_Integer_VALUE else null end) OriginalMoodysScore,
    max(case when field_name = 'OriginalSPRating' then FIELD_String_VALUE else null end) OriginalSPRating,
    max(case when field_name = 'OriginalSPScore' then FIELD_Integer_VALUE else null end) OriginalSPScore,
    max(case when field_name = 'OriginalTrancheBalance' then FIELD_Float_VALUE else null end) OriginalTrancheBalance,
    max(case when field_name = 'OriginalWorstRatingScore' then FIELD_Integer_VALUE else null end) OriginalWorstRatingScore,
    max(case when field_name = 'PaymentPeriod' then FIELD_Integer_VALUE else null end) PaymentPeriod,
    max(case when field_name = 'PercentCCC' then FIELD_Float_VALUE else null end) PercentCCC,
    max(case when field_name = 'PercentCCCLimit' then FIELD_Float_VALUE else null end) PercentCCCLimit,
    max(case when field_name = 'PercentInDefault' then FIELD_Float_VALUE else null end) PercentInDefault,
    max(case when field_name = 'PercentPerforming' then FIELD_Float_VALUE else null end) PercentPerforming,
    max(case when field_name = 'PercentStructured' then FIELD_Float_VALUE else null end) PercentStructured,
    max(case when field_name = 'PrincipalCollectionAccount' then FIELD_Float_VALUE else null end) PrincipalCollectionAccount,
    max(case when field_name = 'PrincipalType' then FIELD_String_VALUE else null end) PrincipalType,
    max(case when field_name = 'ReinvEndDate' then FIELD_Integer_VALUE else null end) ReinvEndDate,
    max(case when field_name = 'ReinvestmentComponent' then FIELD_String_VALUE else null end) ReinvestmentComponent,
    max(case when field_name = 'REIT' then FIELD_Float_VALUE else null end) REIT,
    max(case when field_name = 'RelativeLiquidityScore' then FIELD_Integer_VALUE else null end) RelativeLiquidityScore,
    max(case when field_name = 'ReportedCurrentWAC' then FIELD_Float_VALUE else null end) ReportedCurrentWAC,
    max(case when field_name = 'ReportedCurrentWAS' then FIELD_Float_VALUE else null end) ReportedCurrentWAS,
    max(case when field_name = 'Retail' then FIELD_Float_VALUE else null end) Retail,
    max(case when field_name = 'RiskRetention' then FIELD_String_VALUE else null end) RiskRetention,
    max(case when field_name = 'SecondLienBalance' then FIELD_Float_VALUE else null end) SecondLienBalance,
    max(case when field_name = 'SecondLienBalancePercent' then FIELD_Float_VALUE else null end) SecondLienBalancePercent,
    max(case when field_name = 'SeniorSecuredBalance' then FIELD_Float_VALUE else null end) SeniorSecuredBalance,
    max(case when field_name = 'SeniorSecuredPercent' then FIELD_Float_VALUE else null end) SeniorSecuredPercent,
    max(case when field_name = 'ServicesBusiness' then FIELD_Float_VALUE else null end) ServicesBusiness,
    max(case when field_name = 'ServicesConsumer' then FIELD_Float_VALUE else null end) ServicesConsumer,
    max(case when field_name = 'ShortFall' then FIELD_Float_VALUE else null end) ShortFall,
    max(case when field_name = 'StructuredFinanceObligation' then FIELD_Float_VALUE else null end) StructuredFinanceObligation,
    max(case when field_name = 'Subordination' then FIELD_Float_VALUE else null end) Subordination,
    max(case when field_name = 'TotalContributedBalance' then FIELD_Float_VALUE else null end) TotalContributedBalance,
    max(case when field_name = 'TotalLiability' then FIELD_Float_VALUE else null end) TotalLiability,
    max(case when field_name = 'TotalLoanCount' then FIELD_Integer_VALUE else null end) TotalLoanCount,
    max(case when field_name = 'TrancheCurrency' then FIELD_String_VALUE else null end) TrancheCurrency,
    max(case when field_name = 'TrancheFactor' then FIELD_Float_VALUE else null end) TrancheFactor,
    max(case when field_name = 'TrancheName' then FIELD_String_VALUE else null end) TrancheName,
    max(case when field_name = 'TrancheOrder' then FIELD_String_VALUE else null end) TrancheOrder,
    max(case when field_name = 'TrancheType' then FIELD_String_VALUE else null end) TrancheType,
    max(case when field_name = 'Vintage' then FIELD_Integer_VALUE else null end) Vintage,
    max(case when field_name = 'VolckerCompliant' then FIELD_String_VALUE else null end) VolckerCompliant,
    max(case when field_name = 'WeightedAvgLiability' then FIELD_Float_VALUE else null end) WeightedAvgLiability,
    max(case when field_name = 'AssetContributedBalanceWithBelow70Price' then FIELD_Float_VALUE else null end) Below70, 

    max(case when field_name = 'OvercollaterizationTest' then FIELD_string_VALUE else null end) OCTest, 
    max(case when field_name = 'OvercollaterizationTestCusion' then FIELD_Float_VALUE else null end) OCCushion, 
    max(case when field_name = 'InterestCoverageTest' then FIELD_string_VALUE else null end) ICTest, 
    max(case when field_name = 'InterestDiversionTest' then FIELD_string_VALUE else null end) IDTest,
    max(case when field_name = 'InterestDiversionTestCusion' then FIELD_float_VALUE else null end) IDCushion,

    max(case when field_name = 'HighestClassFailingInterestCoverageTest' then FIELD_string_VALUE else null end) HighestFailIC,
    max(case when field_name = 'HighestClassFailingOvercollaterizationTest' then FIELD_string_VALUE else null end) HighestFailOC,
    max(case when field_name = 'HighestClassFailingOvercollaterizationTestCusion' then FIELD_string_VALUE else null end) HighestFailOCCushion,
    max(case when field_name = 'TrancheRank' then FIELD_integer_VALUE else null end) TrancheRank,
    max(case when field_name = 'WeightedAvgPrice' then FIELD_Float_VALUE else null end) WeightedAvgPrice

    from sec_indicative_info
    group by sec_id) "INFO"
    on pr.sec_id= INFO.sec_id

    left join (select ids1.idvalue as "ID",ids2.idvalue as "TICKER"
    from instruments_ids ids1, instruments_ids ids2
    where ids1.idset=ids2.idset
    and ids2.idtype='BLOOMBERG_TICKER'
    and ids1.idtype in ('ISIN','CUSIP')
    and ids1.validateid='Y'
    and ids2.validateid='Y') "BLOOMBERG"
    on pr.sec_id=BLOOMBERG.ID

    where sm.prc_config_sec_type ='CLO'
    and pr.batch = '"""+batch+"""'

    and pr.market_date = to_date('"""+dat.strftime(formate)+"""','mm/dd/yyyy')
    and sm.sec_status='Approved'
    and pr.sec_id = sm.sec_id
    and pr.sec_id_type = sm.sec_id_type
    and sm.prc_config_key_name in '"""+sectors+"""'

    """

    dfm=pd.read_sql(masterqry,con)

    #re-name to "translated"
    dfm.columns=sql_column_names

    #read colors
    dfc=pd.read_csv(cd)
    dfcc=dfc[['Comments','Px','MESSAGE_ID','BIAS','SOURCE','Cusip','Date2']]

    #find today colors as dfcc1 and otherwise most recent as dfcc2
    dat=dat.strftime(formate)
    dfcc1=dfcc[dfcc.Date2==dat].groupby(['Cusip']).head(1)
    dfcc2=dfcc[dfcc.Date2!=dat].groupby(['Cusip']).head(1).drop(['MESSAGE_ID','Comments','SOURCE',], axis=1)

    dfcc1.columns=['Comments','New Bid Price','Message ID','Bias','Source','Cusip','Quote Date']
    dfcc2.columns=['Last Color','Last Bias','Cusip','Last Date']

    #Leftjoin two color dataframes to Main frame
    dfmf=dfm.merge(dfcc1,on='Cusip',how='left').merge(dfcc2,on='Cusip',how='left')

    #Calculate Bid side of Offer colors
    dfmf['New Bid Price']=np.where(dfmf['Bias']=='OFFER', dfmf['New Bid Price']-dfmf['Price Ask']+dfmf['Price Bid'], dfmf['New Bid Price'])

    #Add in New DM an Diff
    dfmf['New DM']=''
    dfmf['Diff']=dfmf['Price Bid']-dfmf['New Bid Price']

    #Rearrange columns
    #order=['Comments','New Bid Price','Diff','Bias','Quote Date','Message ID','Source','Last Color','Last Bias','Last Date']+column_names
    dfmf=dfmf[all_column_names]

    #Write into Main
    wsm.clear()
    wsm['A1'].options(index=False).value=dfmf

    #Formating
    wsm.autofit(axis="columns")
    #wsm.range(mesid+':'+messid).number_format='#'
    wsm.range(mesid+':'+messid).number_format='#'
    wsm.range("A1:AZ1").api.Font.ColorIndex = 2
    wsm.range("A1:AZ1").color= (31,73,125)  

    #Show numeric message ID instead of scentific notation
    #wsm['D:D'].number_format='0'

    #Closing Connection
    con.close()

def machinelearning():
    style.use('ggplot')
    
    #read master sheet
    currentpath = os.path.dirname(__file__) 
    for filename in os.listdir(currentpath):
        if filename.endswith(".xlsm"):
            fn=filename
            target=currentpath+'\\'+filename

    df = pd.read_excel(target,sheet_name='Main')

    #Set parameter variables
    csp='Cusip'
    bias='Bias'
    quodate='Quote Date'
    dm='New DM'
    nb='New Bid Price'
    pb='Price Bid'
    wal='WAL'
    fm='FloaterMargin'
    mvoc='MarketValueOC'
    occ='OCCushion'
    tier='CollateralManagerTier'
    red='ReinvEndDate'

    colors=['BID','TRADE CONFIRM','BWIIC COVER','BWIC REO','MARKET']
    reliable=['TRADE CONFIRM','BWIIC COVER']
    metric=[dm,wal,fm,mvoc,tier]
    daylag=7  
    peg='peg'
    count='count'

    #Setting date range
    #end=dt.datetime(2022,4,13).date()
    #end=dt.date.today()
    end=df.loc[0,'Date'].date()
    start=end-dt.timedelta(days=daylag)

    #set tool dfs
    df_col = pd.DataFrame()
    df_res = pd.DataFrame()

    #data cleaning
    df[peg]=0
    df[wal]=df[wal].fillna(df[wal].median())
    df[fm]=df[fm].fillna(df[fm].median())
    df[mvoc]=df[mvoc].fillna(df[mvoc].median())
    df[tier]=pd.to_numeric(df[tier].astype("str").str[-1])
    df[tier]=df[tier].fillna(df[tier].median())
    df[quodate]=df[quodate].fillna("1997-04-16").dt.date
    #df[red]=df[quodate].fillna(dt.date.today().year)

    #add helper feathers
    df['rev']=df[red]-dt.datetime.today().year
    df['rev']=df['rev'].fillna(0)

    df['mvocm']=df[mvoc]-df[mvoc].median()
    df['mvocxfm']=df[mvoc]*df[fm]
    df['walxfm']=df[wal]*df[fm]
    metric.extend(['mvocm','mvocxfm','walxfm','rev'])



    #filter colors
    for i in range(df.shape[0]):
        if start<df.loc[i,quodate] and (df.loc[i,bias] in colors):
            df_col=df_col.append(df.loc[i,metric],ignore_index=True)
            if df.loc[i,bias] in reliable:
                df_col=df_col.append(df.loc[i,metric],ignore_index=True)

    metric.insert(1,csp)
    df_res=df.loc[:,metric]

    #import models
    regs=[]
    regs.append(('DeciTree',tree.DecisionTreeRegressor()))
    regs.append(('Linear',linear_model.LinearRegression()))
    regs.append(('AdaBoosterRegressor',AdaBoostRegressor()))
    regs.append(('RandomForestRegressor',RandomForestRegressor(n_estimators=50)))

    #train set
    feat=df_col.iloc[:,1:].to_numpy()
    label=df_col.iloc[:,0].to_numpy()
    met=df_res.iloc[:,2:].to_numpy()
    #train_test_split
    #x_train,x_test,y_train,y_test=train_test_split(feat,label,test_size=0.3)

    #run model
    for Regressor_name,reg in regs:
    
        reg.fit(feat,label)
        res=reg.predict(met)
        df_res[Regressor_name]=res.tolist()

    print(df_res.head(10))
    df_res.to_excel(currentpath+'\\'+'MLDM.xlsx')





